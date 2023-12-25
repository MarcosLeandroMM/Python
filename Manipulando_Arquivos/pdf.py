import PyPDF2
import re
import openpyxl

def extrair_informacoes_caso_ordem_data(texto):
    padrao_caso = re.compile(r'Caso:\s*(\d+)', re.IGNORECASE)
    padrao_ordem_servico = re.compile(r'Ordem de Serviço:\s*(\d+)', re.IGNORECASE)
    padrao_data = re.compile(r'data:\s*(\d{2}/\d{2}/\d{4})', re.IGNORECASE)

    caso_match = padrao_caso.search(texto)
    ordem_servico_match = padrao_ordem_servico.search(texto)
    data_match = padrao_data.search(texto)

    print("Texto da página:", texto)
    print("Match 'Caso':", caso_match.group(1) if caso_match else None)
    print("Match 'Ordem de Serviço':", ordem_servico_match.group(1) if ordem_servico_match else None)
    print("Match 'Data':", data_match.group(1) if data_match else None)

    caso_numero = caso_match.group(1) if caso_match else None
    ordem_servico_numero = ordem_servico_match.group(1) if ordem_servico_match else None
    data = data_match.group(1) if data_match else None

    return caso_numero, ordem_servico_numero, data


def encontrar_palavras_chave_pdf_e_excel(pdf_path, excel_path):
    informacoes_encontradas = []

    with open(pdf_path, 'rb') as file:
        pdf_reader = PyPDF2.PdfReader(file)
        num_paginas = len(pdf_reader.pages)

        for pagina_numero in range(num_paginas):
            pagina = pdf_reader.pages[pagina_numero]
            texto_pagina = pagina.extract_text()

            caso, ordem_servico, data = extrair_informacoes_caso_ordem_data(texto_pagina)

            if caso or ordem_servico or data:
                informacoes_encontradas.append({
                    'Caso': caso,
                    'Ordem de Serviço': ordem_servico,
                    'Data': data
                })

    # Escrever resultados em uma planilha Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Escrever cabeçalho
    cabecalho = ['Caso', 'Ordem de Serviço', 'Data']
    for coluna, titulo in enumerate(cabecalho, start=1):
        sheet.cell(row=1, column=coluna, value=titulo)

    # Escrever dados
    for linha, informacao in enumerate(informacoes_encontradas, start=2):
        for coluna, valor in enumerate(informacao.values(), start=1):
            sheet.cell(row=linha, column=coluna, value=valor)

    # Salvar planilha
    workbook.save(excel_path)

# Exemplo de uso
encontrar_palavras_chave_pdf_e_excel('documento_teste.pdf', 'resultados.xlsx')
